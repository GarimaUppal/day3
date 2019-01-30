from openpyxl import Workbook
import datetime
wb= Workbook()

ws= wb.active
ws['A1'] =42

#adding data in rows
ws.append([1,2,3])

#type conversion
ws['A2']= datetime.datetime.now()

#save the file
wb.save("sample.xlsx")
        

